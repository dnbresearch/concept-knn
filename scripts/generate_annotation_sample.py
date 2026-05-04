"""
Generate a stratified sample of 100 prompts for human annotation.

Usage:
    python generate_annotation_sample.py \
        --sharegpt_data <path_to_sharegpt_processed.json> \
        --lmsys_data <path_to_lmsys_processed.json> \
        --sharegpt_gt <path_to_sharegpt_ground_truth.json> \
        --lmsys_gt <path_to_lmsys_ground_truth.json> \
        --output annotation_sample.xlsx

If your data format is different, adjust the loading functions below.
The script expects:
  - A data file with prompts (keyed by conversation/item ID)
  - A GT file mapping item IDs to coarse/mid/fine categories

Output: An XLSX with two sheets:
  1. "Annotation" - for the human annotator (prompt + blank columns)
  2. "Answer Key" - hidden GT labels for computing agreement later
"""

import json
import random
import argparse
import os

random.seed(42)

# ─── Configuration ───────────────────────────────────────────
SAMPLE_PER_DATASET = 50  # 50 from each dataset = 100 total
MIN_PER_CATEGORY = 1     # At least 1 from each coarse category
MAX_PER_CATEGORY = 5     # Cap per category for balance

# ─── Coarse categories (from the paper's 29 evaluated categories) ───
COARSE_CATEGORIES = [
    "arts & design", "business & finance", "communication & writing",
    "creative writing & fiction", "education & learning", "entertainment & media",
    "food & cooking", "gaming", "general knowledge", "health & medicine",
    "history & culture", "language & translation", "legal",
    "lifestyle & personal", "math & logic", "music",
    "nature & environment", "news & current events", "other",
    "philosophy & ethics", "politics & government", "programming",
    "psychology & relationships", "religion & spirituality",
    "science", "social media", "sports", "technology", "travel & geography"
]


def load_data(data_path, gt_path):
    """
    Load prompts and ground truth. Adjust this function to match your data format.
    
    Expected formats:
    - data: list of dicts with 'id' and 'prompt' (or 'conversation' with first turn)
    - gt: dict mapping id -> {'coarse': str, 'mid': str, 'fine': str}
    
    OR adjust below to match your actual format.
    """
    print(f"Loading data from {data_path}...")
    with open(data_path, 'r') as f:
        data = json.load(f)
    
    print(f"Loading GT from {gt_path}...")
    with open(gt_path, 'r') as f:
        gt = json.load(f)
    
    # Build a dict: id -> prompt text
    prompts = {}
    if isinstance(data, list):
        for item in data:
            item_id = str(item.get('id', item.get('conversation_id', '')))
            # Try different prompt field names
            prompt = item.get('prompt', '')
            if not prompt and 'conversations' in item:
                # ShareGPT format: first human turn
                for turn in item['conversations']:
                    if turn.get('from', turn.get('role', '')) in ('human', 'user'):
                        prompt = turn.get('value', turn.get('content', ''))
                        break
            if not prompt and 'conversation' in item:
                conv = item['conversation']
                if isinstance(conv, list) and len(conv) > 0:
                    prompt = conv[0].get('content', conv[0].get('value', ''))
            if prompt:
                prompts[item_id] = prompt[:500]  # Truncate for readability
    elif isinstance(data, dict):
        for item_id, item in data.items():
            if isinstance(item, str):
                prompts[str(item_id)] = item[:500]
            elif isinstance(item, dict):
                prompt = item.get('prompt', item.get('text', ''))
                prompts[str(item_id)] = prompt[:500]
    
    # Build GT dict: id -> {coarse, mid, fine}
    gt_labels = {}
    if isinstance(gt, dict):
        for item_id, labels in gt.items():
            if isinstance(labels, dict):
                gt_labels[str(item_id)] = labels
            elif isinstance(labels, str):
                gt_labels[str(item_id)] = {'coarse': labels, 'mid': '', 'fine': ''}
    elif isinstance(gt, list):
        for item in gt:
            item_id = str(item.get('id', ''))
            gt_labels[item_id] = {
                'coarse': item.get('coarse', item.get('category_coarse', '')),
                'mid': item.get('mid', item.get('category_mid', '')),
                'fine': item.get('fine', item.get('category_fine', '')),
            }
    
    return prompts, gt_labels


def stratified_sample(prompts, gt_labels, n=50):
    """Stratified sample: proportional to category size, with min/max caps."""
    # Group IDs by coarse category
    by_category = {}
    for item_id, labels in gt_labels.items():
        if item_id not in prompts:
            continue
        coarse = labels.get('coarse', 'unknown').lower().strip()
        if coarse not in by_category:
            by_category[coarse] = []
        by_category[coarse].append(item_id)
    
    # Calculate proportional allocation
    total = sum(len(ids) for ids in by_category.values())
    n_categories = len(by_category)
    
    print(f"  Found {n_categories} categories, {total} items with GT")
    
    # Allocate proportionally with min/max
    allocation = {}
    remaining = n
    for cat, ids in sorted(by_category.items(), key=lambda x: len(x[1])):
        prop = max(MIN_PER_CATEGORY, round(n * len(ids) / total))
        prop = min(prop, MAX_PER_CATEGORY, len(ids))
        allocation[cat] = prop
        remaining -= prop
    
    # Distribute remaining slots to largest categories
    if remaining > 0:
        for cat in sorted(by_category, key=lambda c: len(by_category[c]), reverse=True):
            if remaining <= 0:
                break
            extra = min(remaining, MAX_PER_CATEGORY - allocation.get(cat, 0))
            if extra > 0:
                allocation[cat] = allocation.get(cat, 0) + extra
                remaining -= extra
    
    # Sample from each category
    sample = []
    for cat, count in allocation.items():
        ids = by_category[cat]
        chosen = random.sample(ids, min(count, len(ids)))
        for item_id in chosen:
            sample.append({
                'id': item_id,
                'prompt': prompts[item_id],
                'gt_coarse': gt_labels[item_id].get('coarse', ''),
                'gt_mid': gt_labels[item_id].get('mid', ''),
                'gt_fine': gt_labels[item_id].get('fine', ''),
            })
    
    random.shuffle(sample)  # Shuffle so annotator doesn't see category blocks
    return sample[:n]


def create_xlsx(samples_sg, samples_lm, output_path):
    """Create annotation spreadsheet with two sheets."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("Installing openpyxl...")
        os.system("pip install openpyxl --break-system-packages -q")
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = openpyxl.Workbook()
    
    # ─── Sheet 1: Annotation (what the human sees) ───
    ws = wb.active
    ws.title = "Annotation"
    
    # Styles
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2E4057', end_color='2E4057', fill_type='solid')
    wrap = Alignment(wrap_text=True, vertical='top')
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )
    yellow_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    
    # Headers
    headers = ['#', 'Dataset', 'Prompt (first 500 chars)',
               'Your Coarse Category', 'Your Confidence (1-3)',
               'Notes (optional)']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap
        cell.border = thin_border
    
    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 70
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 25
    
    # Add instructions row
    ws.insert_rows(1)
    ws.merge_cells('A1:F1')
    instr = ws.cell(row=1, column=1)
    instr.value = (
        "INSTRUCTIONS: For each prompt, assign ONE coarse category from the list in the "
        "'Categories' sheet. Rate confidence: 1=unsure, 2=reasonable, 3=certain. "
        "Yellow columns are for your input."
    )
    instr.font = Font(name='Arial', italic=True, size=10, color='333333')
    instr.alignment = wrap
    ws.row_dimensions[1].height = 35
    
    # Data rows
    row = 3  # After header (row 2) 
    all_samples = []
    for s in samples_sg:
        s['dataset'] = 'ShareGPT'
        all_samples.append(s)
    for s in samples_lm:
        s['dataset'] = 'LMSYS'
        all_samples.append(s)
    
    random.shuffle(all_samples)  # Mix datasets
    
    for i, s in enumerate(all_samples, 1):
        ws.cell(row=row, column=1, value=i).border = thin_border
        ws.cell(row=row, column=2, value=s['dataset']).border = thin_border
        
        prompt_cell = ws.cell(row=row, column=3, value=s['prompt'])
        prompt_cell.alignment = wrap
        prompt_cell.border = thin_border
        
        # Yellow fill for annotator columns
        for col in [4, 5, 6]:
            cell = ws.cell(row=row, column=col, value='')
            cell.fill = yellow_fill
            cell.border = thin_border
            cell.alignment = wrap
        
        ws.row_dimensions[row].height = 60
        row += 1
    
    # ─── Sheet 2: Categories reference ───
    ws_cat = wb.create_sheet("Categories")
    ws_cat.cell(row=1, column=1, value="Coarse Category").font = Font(bold=True, name='Arial')
    ws_cat.cell(row=1, column=2, value="Example Topics").font = Font(bold=True, name='Arial')
    ws_cat.column_dimensions['A'].width = 30
    ws_cat.column_dimensions['B'].width = 60
    
    category_examples = {
        "programming": "code, debugging, algorithms, web dev, APIs",
        "creative writing & fiction": "stories, novels, characters, worldbuilding",
        "technology": "hardware, software, gadgets, AI tools",
        "education & learning": "studying, homework, explanations, tutoring",
        "business & finance": "investing, accounting, startups, economics",
        "math & logic": "calculations, proofs, puzzles, statistics",
        "science": "physics, chemistry, biology, research",
        "health & medicine": "symptoms, treatments, nutrition, fitness",
        "communication & writing": "emails, essays, resumes, grammar",
        "entertainment & media": "movies, TV, books, celebrities",
        "gaming": "video games, strategy, game mechanics",
        "language & translation": "translation, language learning, linguistics",
        "arts & design": "drawing, graphic design, architecture",
        "food & cooking": "recipes, cooking techniques, restaurants",
        "legal": "law, contracts, regulations, rights",
        "history & culture": "historical events, civilizations, traditions",
        "philosophy & ethics": "moral questions, logic, existentialism",
        "politics & government": "policy, elections, governance",
        "psychology & relationships": "mental health, dating, social skills",
        "religion & spirituality": "faith, theology, spiritual practices",
        "lifestyle & personal": "productivity, hobbies, self-improvement",
        "music": "instruments, theory, songs, artists",
        "nature & environment": "ecology, climate, animals, geography",
        "travel & geography": "destinations, planning, maps",
        "social media": "platforms, content creation, marketing",
        "sports": "athletics, scores, training",
        "news & current events": "breaking news, journalism, analysis",
        "general knowledge": "trivia, facts, definitions",
        "other": "anything that doesn't fit above",
    }
    
    for i, (cat, examples) in enumerate(sorted(category_examples.items()), 2):
        ws_cat.cell(row=i, column=1, value=cat)
        ws_cat.cell(row=i, column=2, value=examples)
    
    # ─── Sheet 3: Answer Key (hidden — for computing κ later) ───
    ws_key = wb.create_sheet("Answer Key")
    ws_key.cell(row=1, column=1, value="#").font = Font(bold=True)
    ws_key.cell(row=1, column=2, value="ID").font = Font(bold=True)
    ws_key.cell(row=1, column=3, value="Dataset").font = Font(bold=True)
    ws_key.cell(row=1, column=4, value="GT Coarse").font = Font(bold=True)
    ws_key.cell(row=1, column=5, value="GT Mid").font = Font(bold=True)
    ws_key.cell(row=1, column=6, value="GT Fine").font = Font(bold=True)
    
    for i, s in enumerate(all_samples, 1):
        ws_key.cell(row=i+1, column=1, value=i)
        ws_key.cell(row=i+1, column=2, value=s['id'])
        ws_key.cell(row=i+1, column=3, value=s['dataset'])
        ws_key.cell(row=i+1, column=4, value=s['gt_coarse'])
        ws_key.cell(row=i+1, column=5, value=s['gt_mid'])
        ws_key.cell(row=i+1, column=6, value=s['gt_fine'])
    
    ws_key.column_dimensions['B'].width = 20
    ws_key.column_dimensions['D'].width = 25
    ws_key.column_dimensions['E'].width = 35
    ws_key.column_dimensions['F'].width = 40
    
    # Hide answer key sheet
    ws_key.sheet_state = 'hidden'
    
    wb.save(output_path)
    print(f"\nSaved to {output_path}")
    print(f"  - 'Annotation' sheet: {len(all_samples)} prompts for human labeling")
    print(f"  - 'Categories' sheet: {len(category_examples)} category definitions")
    print(f"  - 'Answer Key' sheet: hidden (unhide after annotation to compute κ)")


def main():
    parser = argparse.ArgumentParser(description='Generate stratified annotation sample')
    parser.add_argument('--sharegpt_data', required=True, help='Path to ShareGPT processed data')
    parser.add_argument('--lmsys_data', required=True, help='Path to LMSYS processed data')
    parser.add_argument('--sharegpt_gt', required=True, help='Path to ShareGPT ground truth')
    parser.add_argument('--lmsys_gt', required=True, help='Path to LMSYS ground truth')
    parser.add_argument('--output', default='annotation_sample.xlsx', help='Output XLSX path')
    parser.add_argument('--n_per_dataset', type=int, default=50, help='Samples per dataset')
    args = parser.parse_args()
    
    global SAMPLE_PER_DATASET
    SAMPLE_PER_DATASET = args.n_per_dataset
    
    # Load and sample
    print("=" * 60)
    print("GENERATING STRATIFIED ANNOTATION SAMPLE")
    print("=" * 60)
    
    print("\n--- ShareGPT ---")
    sg_prompts, sg_gt = load_data(args.sharegpt_data, args.sharegpt_gt)
    print(f"  Loaded {len(sg_prompts)} prompts, {len(sg_gt)} GT labels")
    sg_sample = stratified_sample(sg_prompts, sg_gt, n=SAMPLE_PER_DATASET)
    print(f"  Sampled {len(sg_sample)} prompts")
    
    print("\n--- LMSYS ---")
    lm_prompts, lm_gt = load_data(args.lmsys_data, args.lmsys_gt)
    print(f"  Loaded {len(lm_prompts)} prompts, {len(lm_gt)} GT labels")
    lm_sample = stratified_sample(lm_prompts, lm_gt, n=SAMPLE_PER_DATASET)
    print(f"  Sampled {len(lm_sample)} prompts")
    
    # Create XLSX
    print(f"\n--- Creating XLSX ---")
    create_xlsx(sg_sample, lm_sample, args.output)
    
    print("\n" + "=" * 60)
    print("NEXT STEPS:")
    print("1. Open the XLSX and annotate the 'Annotation' sheet")
    print("2. For each prompt, pick ONE coarse category from 'Categories' sheet")
    print("3. Rate confidence: 1=unsure, 2=reasonable, 3=certain")
    print("4. When done, share the completed XLSX back")
    print("5. We'll compute Cohen's κ against the GT and add to the paper")
    print("=" * 60)


if __name__ == '__main__':
    main()
